"""
TikTok Live Finder - Dashboard
==============================
Web dashboard that watches TikTok accounts' Following lists in the background
and shows which followed accounts are LIVE right now.

- Accounts section with full profile cards (avatar, bio, counts, verified)
- Cookie login (Settings) with instant validation + your own profile shown
- Start button per account: scrapes the whole Following list with live
  progress (percent, accounts done, ETA) and streams live discoveries
- Live grid with pulsing indicators, viewer counts, cover thumbnails
- Excel export: <account>_FollowingList_<date>.xlsx with filters + frozen header

Start:            python dashboard.py
Start hidden:     double-click start_dashboard.bat
Then open:        http://<your-pc-ip>:8321  (shown at startup)
"""

import io
import json
import os
import re
import socket
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from hashlib import sha256
from pathlib import Path

import requests
from flask import (Flask, jsonify, make_response, redirect, render_template,
                   request, send_file)

from scraper import UA, check_live, fetch_profile, scrape_following

BASE_DIR = Path(__file__).parent
# DATA_DIR can point at persistent storage (e.g. a Google Drive folder on
# Colab) via the DATA_DIR env var, so scraped lists and resume progress
# survive restarts. Defaults to a local ./data folder.
DATA_DIR = Path(os.environ.get("DATA_DIR") or (BASE_DIR / "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
STATE_FILE = DATA_DIR / "state.json"
COOKIES_FILE = DATA_DIR / "cookies.txt"

PORT = int(os.environ.get("PORT", 8321))
LIVE_WORKERS = 4
LIVE_DELAY = 0.3

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True
state_lock = threading.RLock()
scrape_lock = threading.Lock()


# ----------------------------------------------------------------- state ----

def default_monitor() -> dict:
    return {"profile": None, "secuid": "", "cursor": "0", "cooldown_until": 0,
            "ed_cursor": 0, "ed_page_token": "",
            "following": [], "following_updated": 0, "live": [], "last_scan": 0,
            "scanning": False, "scan_ctl": "", "scrape_ctl": "",
            "loading_following": False,
            "status": "", "progress": {"phase": "", "done": 0, "total": 0, "eta": 0}}


def default_state() -> dict:
    return {"interval": 300, "pin": "", "ed_token": "", "monitors": {},
            "log": [], "self": None}


def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            s = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            for k, v in default_state().items():
                s.setdefault(k, v)
            for m in s["monitors"].values():
                for k, v in default_monitor().items():
                    m.setdefault(k, v)
                m["scanning"] = False
                m["loading_following"] = False
                m["scan_ctl"] = ""
                m["scrape_ctl"] = ""
                m["progress"] = {"phase": "", "done": 0, "total": 0, "eta": 0}
            return s
        except Exception:
            pass
    return default_state()


state = load_state()

# Cloud deployment: env vars override/seed the stored settings so secrets
# don't have to live in the repo. DASHBOARD_PIN also protects a fresh deploy
# from being open to the whole internet.
if os.environ.get("DASHBOARD_PIN"):
    state["pin"] = os.environ["DASHBOARD_PIN"]
if os.environ.get("ED_TOKEN"):
    state["ed_token"] = os.environ["ED_TOKEN"]


def save_state() -> None:
    STATE_FILE.write_text(json.dumps(state, indent=1, ensure_ascii=False),
                          encoding="utf-8")


def add_log(msg: str) -> None:
    with state_lock:
        state["log"].append({"t": time.strftime("%H:%M:%S"), "msg": msg})
        state["log"] = state["log"][-40:]
        save_state()
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


def set_mon(uname: str, **kw) -> None:
    with state_lock:
        mon = state["monitors"].get(uname)
        if mon is not None:
            mon.update(kw)
            save_state()


# --------------------------------------------------------------- cookies ----

def parse_cookies(text: str) -> list[dict]:
    """Accepts a Cookie-Editor JSON export OR a raw 'k=v; k2=v2' header string."""
    text = (text or "").strip()
    if not text:
        return []
    cookies = []
    if text.startswith("["):
        try:
            for c in json.loads(text):
                if c.get("name") and c.get("value") is not None:
                    cookies.append({"name": c["name"], "value": str(c["value"]),
                                    "domain": c.get("domain") or ".tiktok.com",
                                    "path": c.get("path") or "/"})
        except Exception:
            return []
    else:
        for part in text.split(";"):
            if "=" in part:
                k, v = part.split("=", 1)
                if k.strip():
                    cookies.append({"name": k.strip(), "value": v.strip(),
                                    "domain": ".tiktok.com", "path": "/"})
    return cookies


def get_cookies() -> list[dict]:
    if COOKIES_FILE.exists():
        return parse_cookies(COOKIES_FILE.read_text(encoding="utf-8"))
    return []


def cookie_dict() -> dict:
    return {c["name"]: c["value"] for c in get_cookies()}


def session_valid() -> tuple[bool, str]:
    """Ask TikTok whether the saved cookies belong to a live login session."""
    if not get_cookies():
        return False, "no cookies saved yet"
    try:
        s = requests.Session()
        s.headers["User-Agent"] = UA
        for k, v in cookie_dict().items():
            s.cookies.set(k, v, domain=".tiktok.com")
        d = s.get("https://www.tiktok.com/passport/web/account/info/?aid=1988",
                  timeout=15).json()
        data = d.get("data") or {}
        if data.get("user_id_str"):
            return True, data.get("username") or data.get("user_id_str")
        return False, data.get("description") or d.get("message") or "unknown"
    except Exception as e:
        return False, f"check failed: {str(e)[:80]}"


def refresh_self_profile() -> None:
    ok, detail = session_valid()
    with state_lock:
        if not ok:
            state["self"] = {"valid": False, "detail": detail,
                             "checked_at": int(time.time())}
            save_state()
            return
        state["self"] = {"valid": True, "detail": detail,
                         "checked_at": int(time.time()), "profile": None}
        save_state()
    # detail may be an email/handle; try to resolve a display profile
    handle = detail.split("@")[0] if "@" in detail else detail
    prof = fetch_profile(handle, cookie_dict())
    with state_lock:
        if state.get("self"):
            state["self"]["profile"] = prof
            save_state()


# ---------------------------------------------------------------- actions ----

def load_profile_task(uname: str) -> None:
    prof = fetch_profile(uname, cookie_dict())
    set_mon(uname, profile=prof,
            status="" if prof else "could not load profile info")
    if prof:
        add_log(f"@{uname}: profile loaded "
                f"({prof['following']} following, {prof['followers']} followers)")
    else:
        add_log(f"@{uname}: profile info not available")


def start_task(uname: str) -> None:
    """Full pipeline: scrape following list with progress, then live-scan."""
    if not scrape_lock.acquire(blocking=False):
        add_log(f"@{uname}: another following-list load is already running")
        return
    try:
        ok_session, detail = session_valid()
        if not ok_session:
            msg = (f"TikTok cookies problem ({detail}) - log in on tiktok.com "
                   "and re-export cookies in Settings")
            add_log(f"@{uname}: {msg}")
            set_mon(uname, status=msg)
            return

        with state_lock:
            mon = state["monitors"].get(uname)
            if not mon:
                return
            total = (mon.get("profile") or {}).get("following") or 0
            secuid = (mon.get("secuid") or "").strip() \
                or (mon.get("profile") or {}).get("secuid") or ""
            resume_cursor = str(mon.get("cursor") or "0")
            ed_cursor_in = mon.get("ed_cursor") or 0
            ed_token_in = mon.get("ed_page_token") or ""
            # accounts already collected in previous runs - the new batch is
            # added ON TOP of these, so progress must count from here, not 0.
            base_done = len(mon.get("following") or [])
            mon["loading_following"] = True
            mon["scrape_ctl"] = ""                 # clear any old pause/stop
            mon["status"] = ""
            mon["progress"] = {"phase": "following",
                               "done": base_done,
                               "total": total, "eta": 0}
            save_state()
        resume_note = "" if resume_cursor in ("0", "") else " (resuming)"
        add_log(f"@{uname}: START - loading following list{resume_note} "
                f"(logged in as {detail})")

        t0 = time.time()

        def progress(count):
            with state_lock:
                mon = state["monitors"].get(uname)
                if mon is None:
                    return
                p = mon["progress"]
                # cumulative = already-saved accounts + this run's new ones
                done = base_done + count
                p["done"] = done
                if p["total"] and count:
                    rate = count / max(time.time() - t0, 1)
                    p["eta"] = int(max(p["total"] - done, 0) / max(rate, 0.1))

        out = {"ed_cursor_in": ed_cursor_in, "ed_page_token_in": ed_token_in}
        ed_token = (state.get("ed_token") or "").strip()

        def control():
            with state_lock:
                m = state["monitors"].get(uname)
                return (m or {}).get("scrape_ctl", "")

        following, err = scrape_following(uname, get_cookies(), secuid=secuid,
                                          start_cursor=resume_cursor, out=out,
                                          ed_token=ed_token,
                                          progress_cb=progress, log_cb=add_log,
                                          control=control)
        if err and not following:
            # If TikTok rate-limited us, start a cooldown so the user can't
            # deepen the block by pressing Start again immediately.
            cooldown = 0
            if "rate-limit" in err or "empty" in err:
                cooldown = int(time.time()) + 1800   # 30 min
            set_mon(uname, loading_following=False, status=err,
                    cooldown_until=cooldown,
                    progress={"phase": "", "done": 0, "total": 0, "eta": 0})
            add_log(f"@{uname}: {err}")
            return

        # Merge with anything already collected so repeated Starts accumulate
        with state_lock:
            mon = state["monitors"].get(uname)
            existing = {f["username"]: f for f in (mon.get("following") or [])} \
                if mon else {}
        before = len(existing)
        for f in following:
            existing[f["username"]] = {**existing.get(f["username"], {}), **f}
        merged = list(existing.values())
        # Save cursor so the next Start resumes; clear it when the list is done.
        has_more = bool(out.get("has_more"))
        new_cursor = out.get("cursor", "0") if has_more else "0"
        # `err` set with data = partial (e.g. data-API daily limit reached)
        partial_note = err if (err and following) else ""
        # Persist the data-API resume position whenever more accounts remain.
        # EnsembleData paginates by cursor AND/OR page_token, and often returns
        # a null cursor with only a page_token - that token IS the resume key,
        # so keep both. (The old code required a non-null cursor and threw the
        # token away, which made the next Start restart from 0.)
        if has_more:
            ec = out.get("ed_cursor")
            ed_cursor_save = 0 if ec is None else ec
            ed_token_save = out.get("ed_page_token") or ""
        else:
            ed_cursor_save, ed_token_save = 0, ""
        # When we stopped only because of the daily quota, tell the user their
        # progress is saved and Start will continue tomorrow.
        if partial_note and has_more:
            partial_note = (f"{partial_note} - progress saved ({len(merged)} so "
                            f"far); press Start again to continue from here")
        set_mon(uname, following=merged,
                following_updated=int(time.time()),
                loading_following=False, cursor=new_cursor, cooldown_until=0,
                ed_cursor=ed_cursor_save,
                ed_page_token=ed_token_save,
                status=partial_note,
                progress={"phase": "", "done": 0, "total": 0, "eta": 0})
        added = len(merged) - before
        if partial_note:
            add_log(f"@{uname}: PARTIAL - {len(merged)} collected (+{added}); "
                    f"{partial_note}")
        elif has_more:
            add_log(f"@{uname}: {len(merged)} collected so far (+{added} this run) "
                    f"- more remain, press Start again to continue")
        else:
            add_log(f"@{uname}: DONE - {len(merged)} following collected "
                    f"(+{added} this run)")
    finally:
        with state_lock:
            mon = state["monitors"].get(uname)
            if mon is not None:
                mon["loading_following"] = False
                save_state()
        scrape_lock.release()

    scan_task(uname)


# Drop a live from the grid if it hasn't been re-CONFIRMED live within this
# window. Stops "ghost" lives lingering when an account can't be reached
# (rate-limited) for a while - if we can't confirm it's live, we don't claim it.
LIVE_TTL = 600        # seconds (~2 auto-scan cycles)


def _rebuild_live(mon) -> None:
    """Recompute mon['live'] from each account's stored live status/info,
    expiring any live that hasn't been confirmed within LIVE_TTL."""
    now = int(time.time())
    live = []
    for f in mon.get("following") or []:
        if f.get("live") and f.get("live_info"):
            if now - int(f.get("live_checked_at") or 0) > LIVE_TTL:
                f["live"] = False             # stale -> treat as ended
                f.pop("live_info", None)
                continue
            live.append(f["live_info"])
    mon["live"] = live


def scan_task(uname: str) -> None:
    """
    Live-scan every followed account. Supports Pause/Stop via mon['scan_ctl'].
    Does NOT wipe previously-found lives: an account only leaves the live grid
    when it is CONFIRMED offline; a failed/rate-limited check keeps its last
    known status (fixes lives vanishing on a re-scan).
    """
    with state_lock:
        mon = state["monitors"].get(uname)
        if not mon or mon.get("scanning"):
            return
        following = list(mon.get("following") or [])
        if not following:
            return
        mon["scanning"] = True
        mon["scan_ctl"] = ""                   # "", "pause", or "stop"
        _rebuild_live(mon)                      # show existing lives immediately
        mon["progress"] = {"phase": "scan", "done": 0,
                           "total": len(following), "eta": 0}
        save_state()
    try:
        session = requests.Session()
        session.headers["User-Agent"] = UA
        for k, v in cookie_dict().items():
            session.cookies.set(k, v, domain=".tiktok.com")

        names = {f["username"]: f for f in following}
        usernames = list(names)
        t0 = time.time()
        done = 0
        stopped = False

        def one(u):
            r = check_live(u, session)
            time.sleep(LIVE_DELAY)
            return r

        with ThreadPoolExecutor(max_workers=LIVE_WORKERS) as pool:
            i = 0
            while i < len(usernames):
                # ---- pause / stop control (checked between small batches) ----
                with state_lock:
                    ctl = (state["monitors"].get(uname) or {}).get("scan_ctl", "")
                if ctl == "stop":
                    stopped = True
                    break
                while ctl == "pause":
                    time.sleep(0.5)
                    with state_lock:
                        ctl = (state["monitors"].get(uname) or {}).get("scan_ctl", "")
                    if ctl == "stop":
                        stopped = True
                        break
                if stopped:
                    break

                batch = usernames[i:i + LIVE_WORKERS]
                results = [fut.result() for fut in
                           [pool.submit(one, u) for u in batch]]
                i += len(batch)
                done += len(batch)

                with state_lock:
                    mon = state["monitors"].get(uname)
                    if mon is None:
                        return
                    p = mon["progress"]
                    p["done"] = done
                    rate = done / max(time.time() - t0, 1)
                    p["eta"] = int((p["total"] - done) / max(rate, 0.1))
                    for r in results:
                        f = names.get(r["username"])
                        if f is None:
                            continue
                        if not r.get("ok"):
                            continue          # keep previous status on failure
                        f["live_checked_at"] = r.get("checked_at", 0)
                        if r.get("live"):
                            r["display_name"] = f.get("display_name", "")
                            f["live"] = True
                            f["live_info"] = r
                        else:
                            f["live"] = False
                            f.pop("live_info", None)
                    _rebuild_live(mon)         # streams updates to the UI
                    if done % 40 == 0:
                        save_state()

        with state_lock:
            mon = state["monitors"].get(uname)
            if mon is not None:
                mon["last_scan"] = int(time.time())
                mon["scan_ctl"] = ""
                mon["progress"] = {"phase": "", "done": 0, "total": 0, "eta": 0}
                _rebuild_live(mon)
                save_state()
                nlive = len(mon.get("live") or [])
        add_log(f"@{uname}: scan {'stopped' if stopped else 'done'} - "
                f"{nlive} live / {done} checked")
    finally:
        with state_lock:
            mon = state["monitors"].get(uname)
            if mon is not None:
                mon["scanning"] = False
                mon["scan_ctl"] = ""
                save_state()


def worker() -> None:
    while True:
        try:
            with state_lock:
                interval = int(state.get("interval") or 0)
                now = time.time()
                # interval <= 0 disables auto re-scan (manual only)
                due = [] if interval <= 0 else [
                    u for u, m in state["monitors"].items()
                    if m.get("following") and not m.get("scanning")
                    and not m.get("loading_following")
                    and now >= int(m.get("cooldown_until") or 0)
                    and now - m.get("last_scan", 0) >= interval]
            for uname in due:
                scan_task(uname)
        except Exception as e:
            add_log(f"worker error: {str(e)[:100]}")
        time.sleep(5)


# ------------------------------------------------------------------ auth ----

def pin_ok() -> bool:
    pin = state.get("pin") or ""
    if not pin:
        return True
    return request.cookies.get("dash_auth", "") == sha256(pin.encode()).hexdigest()


LOGIN_HTML = """<!doctype html><meta name=viewport content="width=device-width,initial-scale=1">
<style>body{font-family:system-ui;background:#101014;color:#eee;display:flex;
align-items:center;justify-content:center;height:100vh;margin:0}
form{background:#1a1a22;padding:32px;border-radius:14px;text-align:center}
input{font-size:22px;padding:10px;width:140px;text-align:center;border-radius:8px;
border:1px solid #333;background:#101014;color:#fff}
button{margin-top:14px;padding:10px 26px;font-size:16px;border:0;border-radius:8px;
background:#fe2c55;color:#fff}</style>
<form method=post action=/login><h3>Enter PIN</h3>
<input name=pin type=password autofocus><br><button>Open</button></form>"""


@app.route("/login", methods=["POST"])
def login():
    pin = request.form.get("pin", "")
    resp = make_response(redirect("/"))
    if pin and pin == (state.get("pin") or ""):
        resp.set_cookie("dash_auth", sha256(pin.encode()).hexdigest(),
                        max_age=90 * 24 * 3600)
    return resp


# ---------------------------------------------------------------- routes ----

@app.route("/")
def index():
    if not pin_ok():
        return LOGIN_HTML
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    with state_lock:
        out = json.loads(json.dumps(state))
    out["has_cookies"] = bool(get_cookies())
    # keep payload small: cap following preview, UI only needs counts
    for m in out["monitors"].values():
        m["following_count"] = len(m.get("following") or [])
        m.pop("following", None)
    return jsonify(out)


@app.route("/api/add", methods=["POST"])
def api_add():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    u = re.sub(r"[^A-Za-z0-9._]", "", (request.json or {}).get("username", ""))
    if not u:
        return jsonify({"error": "bad username"}), 400
    with state_lock:
        if u not in state["monitors"]:
            state["monitors"][u] = default_monitor()
            state["monitors"][u]["status"] = "loading profile..."
            save_state()
    add_log(f"@{u}: added")
    threading.Thread(target=load_profile_task, args=(u,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/remove", methods=["POST"])
def api_remove():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    u = (request.json or {}).get("username", "")
    with state_lock:
        state["monitors"].pop(u, None)
        save_state()
    add_log(f"@{u}: removed")
    return jsonify({"ok": True})


@app.route("/api/start", methods=["POST"])
def api_start():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    u = (request.json or {}).get("username", "")
    force = bool((request.json or {}).get("force"))
    with state_lock:
        mon = state["monitors"].get(u)
        if mon is None:
            return jsonify({"error": "unknown"}), 400
        cd = int(mon.get("cooldown_until") or 0)
        remaining = cd - int(time.time())
        has_api = bool((state.get("ed_token") or "").strip())
        if remaining > 0 and not force and not has_api:
            mins = (remaining + 59) // 60
            mon["status"] = (f"cooling down — TikTok rate-limited this session. "
                             f"Wait ~{mins} min, then press Start.")
            save_state()
            return jsonify({"cooldown": remaining})
    threading.Thread(target=start_task, args=(u,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/scan", methods=["POST"])
def api_scan():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    u = (request.json or {}).get("username", "")
    if u in state["monitors"]:
        threading.Thread(target=scan_task, args=(u,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/scan_ctl", methods=["POST"])
def api_scan_ctl():
    """Pause / resume / stop a running live scan."""
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    body = request.json or {}
    u = body.get("username", "")
    action = body.get("action", "")           # pause | resume | stop
    with state_lock:
        mon = state["monitors"].get(u)
        if mon is not None and mon.get("scanning"):
            mon["scan_ctl"] = {"pause": "pause", "resume": "",
                               "stop": "stop"}.get(action, "")
            save_state()
    add_log(f"@{u}: scan {action}")
    return jsonify({"ok": True})


@app.route("/api/control", methods=["POST"])
def api_control():
    """Pause / resume / stop whichever operation is running (scrape OR scan)."""
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    body = request.json or {}
    u = body.get("username", "")
    action = body.get("action", "")           # pause | resume | stop
    val = {"pause": "pause", "resume": "", "stop": "stop"}.get(action, "")
    with state_lock:
        mon = state["monitors"].get(u)
        if mon is not None:
            # only one of scrape/scan runs at a time; set both harmlessly
            mon["scan_ctl"] = val
            mon["scrape_ctl"] = val
            save_state()
    add_log(f"@{u}: {action}")
    return jsonify({"ok": True})


@app.route("/api/import", methods=["POST"])
def api_import():
    """Load a following list produced by get_following.js (browser export)."""
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    body = request.json or {}
    u = re.sub(r"[^A-Za-z0-9._]", "", body.get("username", ""))
    raw = body.get("following") or []
    if not u or not isinstance(raw, list) or not raw:
        return jsonify({"error": "file has no username/following data"}), 400

    following = []
    for f in raw:
        if not isinstance(f, dict) or not f.get("username"):
            continue
        following.append({
            "username": f["username"],
            "display_name": f.get("display_name") or f.get("nickname") or "",
            "bio": f.get("bio") or "",
            "verified": bool(f.get("verified")),
            "private": bool(f.get("private")),
            "avatar": f.get("avatar") or "",
            "followers": f.get("followers") or 0,
            "following": f.get("following") or 0,
            "likes": f.get("likes") or 0,
        })
    with state_lock:
        mon = state["monitors"].get(u) or default_monitor()
        # merge with any existing entries
        existing = {x["username"]: x for x in (mon.get("following") or [])}
        for f in following:
            existing[f["username"]] = {**existing.get(f["username"], {}), **f}
        mon["following"] = list(existing.values())
        mon["following_updated"] = int(time.time())
        mon["cooldown_until"] = 0
        mon["status"] = ""
        if not mon.get("profile"):
            mon["profile"] = {"username": u, "nickname": u, "bio": "",
                              "verified": False, "private": False, "avatar": "",
                              "secuid": "", "followers": 0,
                              "following": len(existing), "likes": 0}
        state["monitors"][u] = mon
        save_state()
    add_log(f"@{u}: imported {len(following)} accounts from browser export "
            f"({len(existing)} total)")
    # kick off a live scan right away
    threading.Thread(target=scan_task, args=(u,), daemon=True).start()
    return jsonify({"ok": True, "count": len(existing)})


@app.route("/api/live/<uname>")
def api_live(uname):
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    with state_lock:
        mon = state["monitors"].get(uname) or {}
        return jsonify({"live": mon.get("live") or []})


@app.route("/api/secuid", methods=["POST"])
def api_secuid():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    body = request.json or {}
    u = body.get("username", "")
    val = (body.get("secuid") or "").strip()
    with state_lock:
        mon = state["monitors"].get(u)
        if mon is not None:
            mon["secuid"] = val
            save_state()
    add_log(f"@{u}: secUid {'set' if val else 'cleared'}")
    return jsonify({"ok": True})


@app.route("/api/export_csv/<uname>")
def api_export_csv(uname):
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    import csv as _csv
    with state_lock:
        mon = state["monitors"].get(uname)
        if not mon or not mon.get("following"):
            return jsonify({"error": "no following data - press Start first"}), 400
        following = json.loads(json.dumps(mon["following"]))

    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["Username", "Display Name", "Following", "Followers", "Bio",
                "Live Status", "Visibility", "Profile Link"])
    for f in following:
        if f.get("live"):
            ts = f.get("live_checked_at") or 0
            live_txt = ("Yes (" + datetime.fromtimestamp(ts).strftime(
                "%Y-%m-%d %H:%M") + ")") if ts else "Yes"
        elif f.get("live_checked_at"):
            live_txt = "No"
        else:
            live_txt = "Not checked"
        w.writerow([f.get("username", ""), f.get("display_name", ""),
                    f.get("following", 0), f.get("followers", 0),
                    f.get("bio", ""), live_txt,
                    "Private" if f.get("private") else "Public",
                    f"https://www.tiktok.com/@{f['username']}"])
    data = buf.getvalue().encode("utf-8-sig")   # BOM so Excel reads UTF-8
    fname = f"{uname}_FollowingList_{datetime.now():%Y-%m-%d}.csv"
    add_log(f"@{uname}: exported {len(following)} rows to {fname}")
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype="text/csv")


@app.route("/api/export/<uname>")
def api_export(uname):
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    with state_lock:
        mon = state["monitors"].get(uname)
        if not mon or not mon.get("following"):
            return jsonify({"error": "no following data - press Start first"}), 400
        following = json.loads(json.dumps(mon["following"]))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Following"
    headers = ["Username", "Display Name", "Following", "Followers", "Bio",
               "Live Status", "Visibility", "Profile Link"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="FE2C55")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill

    link_font = Font(color="0563C1", underline="single")
    for f in following:
        if f.get("live"):
            ts = f.get("live_checked_at") or 0
            live_txt = "Yes (" + datetime.fromtimestamp(ts).strftime(
                "%Y-%m-%d %H:%M") + ")" if ts else "Yes"
        elif f.get("live_checked_at"):
            live_txt = "No"
        else:
            live_txt = "Not checked"
        url = f"https://www.tiktok.com/@{f['username']}"
        ws.append([f.get("username", ""), f.get("display_name", ""),
                   f.get("following", 0), f.get("followers", 0),
                   f.get("bio", ""), live_txt,
                   "Private" if f.get("private") else "Public", url])
        cell = ws.cell(row=ws.max_row, column=8)
        cell.hyperlink = url
        cell.font = link_font

    widths = [22, 24, 12, 12, 40, 22, 11, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{uname}_FollowingList_{datetime.now():%Y-%m-%d}.xlsx"
    add_log(f"@{uname}: exported {len(following)} rows to {fname}")
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument"
                 ".spreadsheetml.sheet")


@app.route("/api/settings", methods=["POST"])
def api_settings():
    if not pin_ok():
        return jsonify({"error": "auth"}), 401
    body = request.json or {}
    cookies_text = (body.get("cookies") or "").strip()
    pin_changed = False
    with state_lock:
        state["interval"] = max(60, int(body.get("interval") or 300))
        if "ed_token" in body:
            state["ed_token"] = (body.get("ed_token") or "").strip()
        new_pin = body.get("pin", state.get("pin") or "")
        if new_pin != state.get("pin"):
            state["pin"] = new_pin
            pin_changed = True
        save_state()
    if cookies_text:
        parsed = parse_cookies(cookies_text)
        if parsed:
            COOKIES_FILE.write_text(cookies_text, encoding="utf-8")
            threading.Thread(target=_cookie_check_task, daemon=True).start()
        else:
            add_log("cookies NOT saved - could not parse the pasted text")
    resp = make_response(jsonify({"ok": True}))
    if pin_changed and state["pin"]:
        resp.set_cookie("dash_auth", sha256(state["pin"].encode()).hexdigest(),
                        max_age=90 * 24 * 3600)
    return resp


def _cookie_check_task() -> None:
    refresh_self_profile()
    with state_lock:
        me = state.get("self") or {}
    if me.get("valid"):
        add_log(f"cookies saved - LOGIN OK (logged in as {me.get('detail')})")
    else:
        add_log(f"cookies saved but LOGIN NOT WORKING ({me.get('detail')}). "
                "Log in on tiktok.com, refresh, export cookies again.")


def lan_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


if __name__ == "__main__":
    threading.Thread(target=worker, daemon=True).start()
    threading.Thread(target=refresh_self_profile, daemon=True).start()
    print("=" * 52)
    print("  TikTok Live Finder dashboard is running")
    print(f"    On this PC:    http://localhost:{PORT}")
    print(f"    On your phone: http://{lan_ip()}:{PORT}   (same Wi-Fi)")
    print("=" * 52)
    app.run(host="0.0.0.0", port=PORT, debug=False)
